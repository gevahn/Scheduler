import numpy as np
import openpyxl as xl
from datetime import time
from profParser import *


studPrefMatrix = np.array([1,1])
students = []
numberOfStudents = 0

def parseStuds():
	global students
	global studPrefMatrix
	global numberOfStudents

	parseProfs()

	wb = xl.load_workbook("students.xlsx")
	ws = wb.active

	studPrefMatrix = np.zeros((ws.max_row - 1, len(profDict)))
	students = []
	numberOfStudents = ws.max_row - 1

	for row in range(2,ws.max_row + 1):
		name =  ws.cell(column=1, row=row).value
		students.append(name)
		for col in range(3,11):
			preference = ws.cell(column=col, row=row).value
			if preference != None:
				studPrefMatrix[row - 2, profDict[preference]] = 11 - col 
		totalPoints = 36
		studPoints = np.sum(studPrefMatrix[row - 2, :])
		studPrefMatrix[row - 2, :] = studPrefMatrix[row - 2, :] * (totalPoints / studPoints) 

	print np.sum(studPrefMatrix, axis=0)
	print np.sum((np.sum(studPrefMatrix, axis=0) - np.sum(getProfTimeMatrix(),axis=1)) [(np.sum(studPrefMatrix, axis=0) - np.sum(getProfTimeMatrix(),axis=1)) > 0])
	print studPrefMatrix.shape
	print profDict

def getNumberOfStudents():
	return numberOfStudents

def getStudPrefMatrix():
	return studPrefMatrix

def getStudents():
	return students

if __name__== "__main__":
	parseStuds()
