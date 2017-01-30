import numpy as np
import openpyxl as xl
import datetime as dt

profDict = {}
profTimeMatrix =np.array([1,1])
numberOfProfs = 27

meetingTimes = {dt.time(10,30):0 ,dt.time(11,10):1 ,dt.time(11,50):2,dt.time(15,00):3,dt.time(15,40):4,dt.time(16,20):5,dt.time(17,00):6,dt.time(17,30):7}
numberOfTimeSlots = len(meetingTimes)

def parseProfs():
	global profDict
	global profTimeMatrix 
	global numberOfProfs

	wb = xl.load_workbook("profs.xlsx")
	ws = wb.active

	profTimeMatrix = np.zeros((ws.max_row,ws.max_column - 2))
	numberOfProfs = ws.max_row

	for row in range(1,ws.max_row + 1):
		name = ws.cell(column=1, row=row).value
		profDict[name] = row - 1
		meetingSize = ws.cell(column=2, row=row).value
		for col in range(3,ws.max_column + 1):
			time = ws.cell(column=col, row=row).value
			if time != None:
				profTimeMatrix[row-1, meetingTimes[time]] = meetingSize

	print np.sum(profTimeMatrix,axis=1)


def getNumberOfProfs():
	return numberOfProfs

def getProfDict():
	return profDict

def getProfTimeMatrix():
	return profTimeMatrix

def getNumberOfTimeSlots():
	return numberOfTimeSlots

def getMeetingTimes():
	return meetingTimes
