import numpy as np
import openpyxl as xl
from profParser import *
from studentParser import *

parseStuds()


def move(matrix):
	meetings = np.argwhere(matrix == 1)
	randomMeeting = np.random.randint(meetings.shape[0])
	selectedMeeting = meetings[randomMeeting]
	matrixToReturn = np.copy(matrix)
	matrixToReturn[tuple(selectedMeeting)] = 0
	randomNewTime = np.random.randint(numberOfTimeSlots + 1)
	matrixToReturn[selectedMeeting[0],selectedMeeting[1],randomNewTime] = 1
	return matrixToReturn
	
def energy(matrix, verbose = False):
	studPrefCrit = np.sum(np.sum(matrix,axis=2) * getStudPrefMatrix())
	meetingOverlapPenelty = np.sum((np.sum(matrix,axis=1)) > 1)
	overflowMatrix = np.sum(matrix,axis=0) - getProfTimeMatrix() 
	meetingOverflowPenelty = np.sum(overflowMatrix[overflowMatrix > 0])

	if verbose:
		print "Student preference value: %f"%(studPrefCrit)
		print "Number Of Overlaps: %d"%(meetingOverlapPenelty)
		print "Number Of Overflows: %d"%(meetingOverflowPenelty)
	return -studPrefCrit + meetingOverlapPenelty * 200 + meetingOverflowPenelty * 2000


masterMatrix = np.zeros((getNumberOfStudents(), getNumberOfProfs(), getNumberOfTimeSlots() + 1))

numberOfTimeSlots = getNumberOfTimeSlots()
for i in range(getNumberOfStudents()):
	for j in range(getNumberOfProfs()):
		if getStudPrefMatrix()[i,j] != 0:
			masterMatrix[i,j,numberOfTimeSlots] = 1

initialTemp = 10000.0
temperature = initialTemp
numberOfSteps = 1000000
currentEnergy = energy(masterMatrix[:,:,:-1])
bestEnergy = 0
bestMatrix = np.copy(masterMatrix)

for i in range(numberOfSteps):
	if i % 100000 == 0:
		print "Energy: %f"%(currentEnergy)
		print "Temp: %f"%(temperature)
		print "Step: %d"%(i)
		print "Unallocated: %d"%(np.sum(masterMatrix[:,:,-1:]))
	newMatrix = move(masterMatrix)
	newEnergy = energy(newMatrix[:,:,:-1])
	if np.exp(-(newEnergy - currentEnergy) / temperature) > np.random.rand():
		masterMatrix = newMatrix
		currentEnergy = newEnergy

	if bestEnergy > currentEnergy:
		bestMatrix = np.copy(masterMatrix)

	temperature = initialTemp / pow(i+ 100, 5.0 / 7)

print "SA Summary:"
print "Energy: %f"%(bestEnergy)
print "Unallocated: %d"%(np.sum(masterMatrix[:,:,-1:]))
energy(masterMatrix[:,:,:-1], True)
print "-----------------------------------------"

wb = xl.Workbook()

ws = wb.active
for dimension in ws.column_dimensions.values():
    dimension.auto_size = True

meetingTimes = getMeetingTimes()

ws['A1'] = "First Name"
ws['B1'] = "Last Name"

col = 3
for time in sorted(meetingTimes, key=meetingTimes.__getitem__):
	ws.cell(column=col, row = 1, value = time)
	col = col + 1	


studNames = getStudents()
profDict= getProfDict()
profNames = sorted(profDict, key=profDict.__getitem__)
for row in range(bestMatrix.shape[0]):
	for col in range(bestMatrix.shape[1]):
		firstName = (studNames[row].split(' '))[0]
		ws.cell(column=1, row = 2 + row, value = firstName)
		lastName = studNames[row].split(' ')[1]
		ws.cell(column=2, row = 2 + row, value = lastName)
		ws.cell(column=2, row = 2 + row, value = lastName)
		for time in range(bestMatrix.shape[2] - 1):
			if bestMatrix[row,col,time] == 1:
				ws.cell(column=3 + time, row = 2 + row, value = profNames[col]) 

			
dims = {}
for r in ws.rows:
    for c in r:
        if isinstance(c.value, basestring):
            dims[c.column] = max((dims.get(c.column, 0), len(c.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value

wb.save('studentSchedule.xlsx')

'----------------------'

wb = xl.Workbook()

ws = wb.active
ws['A1'] = "Faculty"
ws['B1'] = "Location"

col = 3
for time in sorted(meetingTimes, key=meetingTimes.__getitem__):
        ws.cell(column=col, row = 1, value = time)
        col = col + 1



currentRow = 2
studentsInMeeting = currentRow
for profIdx in range(bestMatrix.shape[1]):
	profName = profNames[profIdx]
	for timeIdx in range(bestMatrix.shape[2] - 1):
		studentsInMeeting = currentRow
		for studIdx in range(bestMatrix.shape[0]):
			if bestMatrix[studIdx, profIdx, timeIdx] == 1:
				ws.cell(column=1, row = studentsInMeeting, value = profName)
				ws.cell(column=3 + timeIdx, row = studentsInMeeting, value = studNames[studIdx]) 
				studentsInMeeting = studentsInMeeting + 1	
	currentRow = currentRow  + int(np.max(np.sum(bestMatrix[:,:,:-1], axis=0)[profIdx,:]))

dims = {}
for r in ws.rows:
    for c in r:
        if isinstance(c.value, basestring):
            dims[c.column] = max((dims.get(c.column, 0), len(c.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value


wb.save('profSchedule.xlsx')


'-----------------'

wb = xl.Workbook()
ws = wb.active

ws['A1'] = "Faculty"
ws['B1'] = "# of slots"
ws['C1'] = "# of students"
ws['D1'] = "# of unschedulable"


profPref = np.sum(getProfTimeMatrix(), axis = 1)
studPref = np.sum(getStudPrefMatrix() > 0, axis = 0)
for profIdx in range(len(profPref)):
        profName = profNames[profIdx]
	ws.cell(column=1, row = profIdx + 2, value = profName)
	ws.cell(column=2, row = profIdx + 2, value = profPref[profIdx])
	ws.cell(column=3, row = profIdx + 2, value = studPref[profIdx])
	ws.cell(column=4, row = profIdx + 2, value = max(0, studPref[profIdx] - profPref[profIdx]))

dims = {}
for r in ws.rows:
    for c in r:
        if isinstance(c.value, basestring):
            dims[c.column] = max((dims.get(c.column, 0), len(c.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value



wb.save('report.xlsx')
